# 导入相关的库
from PIL import Image
import pytesseract
import cv2 as cv
import numpy as np
import cv2
import os
import re
import pandas as pd
import openpyxl


# binary
def threshold_custome(image):
    gray = cv.cvtColor(image, cv.COLOR_BGR2GRAY)
    h, w = gray.shape[:2]  # 去shape中的前两个通道
    m = np.reshape(gray, [1, w * h])  # reshape()函数重新定义了原张量的阶数,弄成了1行w*h列的一个向量了。
    mean = m.sum() / (w * h)  # 求出整个灰度图像的平均值
    ret, binary = cv.threshold(gray, mean, 255, cv.THRESH_BINARY)
    # 这个函数的第一个参数就是原图像，原图像应该是灰度图。
    # 第二个参数就是用来对像素值进行分类的阈值。
    # 第三个参数就是当像素值高于（有时是小于）阈值时应该被赋予的新的像素值
    # 第四个参数来决定阈值方法
    return binary



# 遍历识别文字
files_path = r'C:\Users\12768\Desktop\Attachment\AttachLittle'
file = os.listdir(files_path)
Timelist = []
temper1 = []
temper2 = []
for f in file:
    real_url = os.path.join(files_path, f)
    # 读取图片
    img = cv2.imread(real_url)
    # # 二值化
    # img_binary = threshold_custome(img)
    # 切割
    img_little = img[:70, :110]
    # 文字识别
    text = pytesseract.image_to_string(img_little, lang="chi_sim")
    print(text)
    # 正则表达式
    number = re.findall("\d+", text)  # 输出结果为列表
    print(number)
    # print(number[0])
    # 将文字添加到li
    # Timelist.append(number[0])
    # temper1.append(number[2])
    # temper2.append(number[-1])

print(Timelist)
print(temper1)
print(temper2)

# 插入excel指定位置
'''
distance_list是一个列表，我们的目标是将该列表作为一列插入表格
'''
# 先打开我们的目标表格，再打开我们的目标表单
wb = openpyxl.load_workbook(r'C:\Users\12768\Desktop\A1.xlsx')
ws = wb['Sheet1']
# 取出distance_list列表中的每一个元素，openpyxl的行列号是从1开始取得，所以我这里i从1开始取
for i in range(1, len(Timelist) + 1):
    a = Timelist[i - 1]
    b = temper1[i - 1]
    c = temper2[i - 1]
    # 写入位置的行列号可以任意改变，这里我是从第2行开始按行依次插入第11列
    ws.cell(row=i + 42, column=2).value = a
    ws.cell(row=i + 42, column=3).value = b
    ws.cell(row=i + 42, column=4).value = c

# 保存操作
wb.save(r'C:\Users\12768\Desktop\Attachment22.xlsx')
